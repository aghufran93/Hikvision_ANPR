#!/usr/bin/env python3
"""
Hikvision ANPR Dashboard v2.0.

The v1 dashboard called glob.glob() + open()+json.load() on every JSON
file in output/Entry and output/Exit, on every single /api/events
request. That doesn't scale: at current volume (~240 events/day) it
gets a little slower every day, forever.

v2.0 keeps an in-memory index (EventIndex) that a single background
thread updates incrementally - each scan cycle only stats directory
entries (cheap) and only opens+parses a JSON file if it's new or its
mtime changed since the last scan. Request handlers never touch disk;
they read the already-built, already-sorted in-memory list. A small
SSE endpoint (/api/events/stream) pushes a lightweight "something
changed, go refetch" signal so the frontend updates immediately
instead of waiting for its next poll tick.

/api/events response shape is unchanged from v1 (aside from the
additive "version" field), so nothing downstream breaks.
"""

import io
import os
import json
import threading
import time
from datetime import datetime, time as dtime

from flask import (
    Flask,
    render_template,
    jsonify,
    request,
    send_file,
    send_from_directory,
    abort,
    Response,
    stream_with_context,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = "/home/rixos/Hikvision_ANPR"

ENTRY_DIR = os.path.join(BASE_DIR, "output", "Entry")
EXIT_DIR = os.path.join(BASE_DIR, "output", "Exit")

MAX_EVENTS_RETURNED = 500

# Excel/PDF export embeds a thumbnail per row, so unlike /api/events (a
# plain JSON list) an unbounded export could mean generating a
# multi-hundred-MB file. Callers must narrow their filters instead of
# silently getting a truncated report.
MAX_EXPORT_ROWS = 2000
EXPORT_THUMB_PX = 80

# Camera considered "live" if the listener process saw stream data
# (a chunk on the multipart HTTP connection - actual vehicle events
# or the camera's own periodic heartbeat XML, whichever comes first)
# within this many seconds. Generous on purpose: this is a
# connectivity check, not a vehicle-throughput check, so it shouldn't
# flip to "stale" during a quiet period with no cars.
CAMERA_STATUS_STALE_SECONDS = float(
    os.getenv("DASHBOARD_CAMERA_STATUS_STALE_SECONDS", "60")
)

# How often the background thread re-checks the output directories for
# new/modified files. This is not the browser's refresh rate - the
# browser is pushed an update immediately via SSE when something
# actually changes; this interval just bounds the worst-case delay
# between a file landing on disk and the dashboard noticing it.
INDEX_SCAN_INTERVAL = float(os.getenv("DASHBOARD_INDEX_SCAN_INTERVAL_S", "0.3"))

# SSE heartbeat so idle connections aren't silently dropped by
# intermediate proxies/load balancers.
SSE_HEARTBEAT_SECONDS = 15

# Client-side polling fallback interval (used alongside SSE as a
# robustness safety net, and as the sole refresh mechanism for any
# client where SSE is blocked). Configurable via .env; 250ms default
# per spec - deliberately not lower, since that would just add load
# for no perceptible benefit given ANPR events are discrete.
try:
    DASHBOARD_POLL_INTERVAL_MS = int(os.getenv("DASHBOARD_POLL_INTERVAL_MS", "250"))
except ValueError:
    DASHBOARD_POLL_INTERVAL_MS = 250

app = Flask(__name__)


# ============================================================
# EVENT SHAPING HELPERS
# (unchanged from v1 - the JSON-on-disk schema and the API response
# shape built from it are both preserved for backward compatibility)
# ============================================================

def safe_load_json(filename):
    try:
        with open(filename, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        app.logger.warning("Unable to read %s: %s", filename, exc)
        return None


def format_event_time(value):
    if not value:
        return {"display": "-", "date": "-", "time": "-"}

    try:
        dt = datetime.fromisoformat(value)
        return {
            "display": dt.strftime("%d/%m/%Y %H:%M:%S"),
            "date": dt.strftime("%d/%m/%Y"),
            "time": dt.strftime("%H:%M:%S"),
        }
    except Exception:
        return {"display": value, "date": "", "time": ""}


def normalise_location(value):
    if not value:
        return "-"

    mapping = {
        "DXB": "Dubai",
        "AUH": "Abu Dhabi",
        "SHJ": "Sharjah",
        "AJM": "Ajman",
        "UAQ": "Umm Al Quwain",
        "RAK": "Ras Al Khaimah",
        "FUJ": "Fujairah",
    }

    return mapping.get(value.upper(), value)


def build_event(data, direction, json_file):
    anpr = data.get("anpr", {})
    camera = data.get("camera", {})
    images = data.get("images", {})

    lpr_read = anpr.get("lpr_read", True)
    full_plate = anpr.get("full_plate")

    if not lpr_read or not full_plate:
        full_plate = "LPR not Read"
        lpr_read = False

    country_code = anpr.get("country") or anpr.get("area") or ""
    province = anpr.get("province") or ""
    location_code = province or country_code

    event_time = format_event_time(data.get("event_time"))

    confidence = anpr.get("confidence")

    if not lpr_read:
        confidence_display = "-"
    elif confidence:
        confidence_display = f"{confidence}%"
    else:
        confidence_display = "-"

    return {
        "id": os.path.basename(json_file).replace(".json", ""),
        "direction": direction,
        "event_time": data.get("event_time", ""),
        "display_time": event_time["display"],
        "date": event_time["date"],
        "time": event_time["time"],
        "lpr_read": lpr_read,
        "plate_number": anpr.get("plate_number", ""),
        "category": anpr.get("category", ""),
        "full_plate": full_plate,
        "country": country_code,
        "state": normalise_location(location_code),
        "confidence": confidence,
        "confidence_display": confidence_display,
        "plate_color": anpr.get("plate_color") or "-",
        "plate_type": anpr.get("plate_type") or "-",
        "vehicle_type": anpr.get("vehicle_type") or "-",
        "vehicle_color": anpr.get("vehicle_color") or "-",
        "movement": anpr.get("direction") or "-",
        "camera_name": camera.get("name") or direction,
        "plate_image": images.get("plate"),
        "vehicle_image": images.get("vehicle"),
        "full_image": images.get("full"),
    }


def build_stats(events):
    today = datetime.now().strftime("%d/%m/%Y")

    today_events = [event for event in events if event["date"] == today]

    entries = sum(1 for event in today_events if event["direction"] == "Entry")
    exits = sum(1 for event in today_events if event["direction"] == "Exit")
    read_events = sum(1 for event in today_events if event["lpr_read"])
    total = len(today_events)

    read_rate = round((read_events / total) * 100, 1) if total else 0

    return {
        "total": total,
        "entry": entries,
        "exit": exits,
        "read": read_events,
        "not_read": total - read_events,
        "read_rate": read_rate,
    }


# ============================================================
# FILTERING (shared by /api/events/query and /api/export)
# ============================================================

def parse_date_param(value):
    if not value:
        return None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_time_param(value):
    if not value:
        return None

    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def parse_datetime_bound(date_str, time_str, end_of_day):
    """Combines a date param with an optional time-of-day param into a
    single bound for range filtering. Without a time, the bound covers
    the whole day (00:00 for a start bound, 23:59:59.999999 for an end
    bound) so plain date filtering keeps working unchanged."""

    date_value = parse_date_param(date_str)

    if date_value is None:
        return None

    time_value = parse_time_param(time_str)

    if time_value is None:
        time_value = dtime(23, 59, 59, 999999) if end_of_day else dtime(0, 0)

    return datetime.combine(date_value, time_value)


def parse_event_datetime(event_time):
    """Event timestamps are stored tz-aware (camera's local offset);
    filter bounds typed into the date/time inputs are naive local wall
    clock values with no offset, so comparison drops tzinfo rather than
    assuming UTC."""

    if not event_time:
        return None

    try:
        return datetime.fromisoformat(event_time).replace(tzinfo=None)
    except ValueError:
        return None


def apply_filters(events, args):
    start_dt = parse_datetime_bound(
        args.get("start_date"), args.get("start_time"), end_of_day=False
    )
    end_dt = parse_datetime_bound(
        args.get("end_date"), args.get("end_time"), end_of_day=True
    )
    direction = args.get("direction") or ""
    lpr = args.get("lpr") or ""
    vehicle_type = args.get("vehicle_type") or ""
    plate_type = args.get("plate_type") or ""
    search = (args.get("search") or "").strip().lower()

    def matches(event):
        if direction and event["direction"] != direction:
            return False

        if lpr == "read" and not event["lpr_read"]:
            return False

        if lpr == "not-read" and event["lpr_read"]:
            return False

        if vehicle_type and event["vehicle_type"] != vehicle_type:
            return False

        if plate_type and event["plate_type"] != plate_type:
            return False

        if search and search not in event["full_plate"].lower():
            return False

        if start_dt or end_dt:
            event_dt = parse_event_datetime(event["event_time"])

            if event_dt is None:
                return False
            if start_dt and event_dt < start_dt:
                return False
            if end_dt and event_dt > end_dt:
                return False

        return True

    return [event for event in events if matches(event)]


def read_camera_status(role):
    """Reads the status.json a listener process (anpr_common.py's
    EventCache) writes on every watchdog tick (~1s) and on every
    connect/disconnect. Lets the dashboard show camera connectivity
    right after a restart instead of waiting for an actual vehicle
    event to prove the stream is alive."""

    path = os.path.join(BASE_DIR, "xml_temp", role, "status.json")

    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {
            "role": role,
            "online": False,
            "connected": False,
            "last_frame_seconds_ago": None,
        }

    last_frame_at = data.get("last_frame_at")
    last_frame_seconds_ago = (
        time.time() - last_frame_at if last_frame_at else None
    )

    online = bool(data.get("connected")) and (
        last_frame_seconds_ago is not None
        and last_frame_seconds_ago < CAMERA_STATUS_STALE_SECONDS
    )

    return {
        "role": role,
        "online": online,
        "connected": bool(data.get("connected")),
        "last_frame_seconds_ago": last_frame_seconds_ago,
    }


def resolve_image_path(event, image_field):
    filename = event.get(image_field)

    if not filename:
        return None

    directory = ENTRY_DIR if event["direction"] == "Entry" else EXIT_DIR
    path = os.path.join(directory, filename)

    return path if os.path.isfile(path) else None


def make_thumbnail(path, max_px):
    """Downscaled JPEG bytes for embedding in an export. The captured
    plate crops are full-resolution camera JPEGs (hundreds of KB each);
    embedding those directly per row would turn a few hundred events
    into a multi-hundred-MB file, so this re-encodes a small thumbnail
    instead of relying on the spreadsheet/PDF viewer's display scaling,
    which only resizes on screen and does not shrink the stored bytes."""

    try:
        with PILImage.open(path) as src:
            thumb = src.convert("RGB")
            thumb.thumbnail((max_px, max_px))

            buf = io.BytesIO()
            thumb.save(buf, format="JPEG", quality=72)
            buf.seek(0)

            return buf
    except Exception:
        app.logger.warning("Failed to build thumbnail for %s", path)
        return None


# ============================================================
# EVENT INDEX
#
# One in-memory copy of every parsed event, keyed by its source JSON
# path, kept current by a single background thread. Request handlers
# only ever read from this - never touch disk.
# ============================================================

class EventIndex:

    def __init__(self):
        self.lock = threading.Lock()
        self._mtimes = {}   # path -> mtime last parsed
        self._events = {}   # path -> built event dict

        self.version = 0
        self._sorted_cache = None
        self._sorted_cache_version = -1

        self._subscribers = []
        self._sub_lock = threading.Lock()

    def scan_once(self):
        """Stat every file in both output directories (cheap); only
        open+parse the ones that are new or whose mtime changed since
        the last scan. Returns True if the index actually changed."""

        changed = False

        for directory, direction in ((ENTRY_DIR, "Entry"), (EXIT_DIR, "Exit")):
            try:
                with os.scandir(directory) as it:
                    entries = [e for e in it if e.name.endswith(".json")]
            except FileNotFoundError:
                continue

            seen_paths = set()

            for entry in entries:
                path = entry.path
                seen_paths.add(path)

                try:
                    mtime = entry.stat().st_mtime
                except OSError:
                    continue

                with self.lock:
                    known_mtime = self._mtimes.get(path)

                if known_mtime == mtime:
                    continue

                data = safe_load_json(path)

                if not data:
                    continue

                try:
                    event = build_event(data, direction, path)
                except Exception as exc:
                    app.logger.warning("Error processing %s: %s", path, exc)
                    continue

                with self.lock:
                    self._events[path] = event
                    self._mtimes[path] = mtime

                changed = True

            with self.lock:
                known_for_dir = [
                    p for p in self._mtimes
                    if os.path.dirname(p) == directory
                ]

            for path in known_for_dir:
                if path not in seen_paths:
                    with self.lock:
                        self._mtimes.pop(path, None)
                        self._events.pop(path, None)
                    changed = True

        if changed:
            with self.lock:
                self.version += 1
            self._notify_subscribers()

        return changed

    def get_sorted_events(self):
        with self.lock:
            if (
                self._sorted_cache is not None
                and self._sorted_cache_version == self.version
            ):
                return self._sorted_cache, self.version

            # Never sort by filename - always by the camera's own
            # event timestamp (ISO8601, consistent timezone offset
            # across both cameras, so lexicographic sort is correct).
            events = sorted(
                self._events.values(),
                key=lambda event: event.get("event_time", ""),
                reverse=True,
            )

            self._sorted_cache = events
            self._sorted_cache_version = self.version

            return events, self.version

    # ---- SSE subscriber management ----

    def subscribe(self):
        wake = threading.Event()

        with self._sub_lock:
            self._subscribers.append(wake)

        return wake

    def unsubscribe(self, wake):
        with self._sub_lock:
            if wake in self._subscribers:
                self._subscribers.remove(wake)

    def _notify_subscribers(self):
        with self._sub_lock:
            subscribers = list(self._subscribers)

        for wake in subscribers:
            wake.set()


index = EventIndex()


def _background_scanner():
    while True:
        try:
            index.scan_once()
        except Exception:
            app.logger.exception("Event index scan failed")

        time.sleep(INDEX_SCAN_INTERVAL)


# Build the index once, synchronously, before serving any request -
# otherwise the first requests after a restart would see an empty
# dashboard while the background thread catches up.
index.scan_once()

_scanner_thread = threading.Thread(
    target=_background_scanner,
    daemon=True,
    name="dashboard-index-scanner",
)
_scanner_thread.start()


# ============================================================
# EXPORT (Excel / PDF)
#
# Filters are applied by the caller (apply_filters) before either of
# these is called - both just render whatever event list they're given.
# Thumbnails are embedded per the "embed plate thumbnail per row"
# requirement; MAX_EXPORT_ROWS in /api/export is what keeps that from
# turning into an unbounded-size file.
# ============================================================

EXPORT_COLUMNS = [
    ("date", "Date"),
    ("time", "Time"),
    ("direction", "Movement"),
    ("full_plate", "Plate Number"),
    ("state", "State/Emirate"),
    ("vehicle_type", "Vehicle Type"),
    ("plate_type", "Plate Type"),
    ("vehicle_color", "Vehicle Colour"),
    ("plate_color", "Plate Colour"),
    ("confidence_display", "Confidence"),
    ("camera_name", "Camera"),
]


def lpr_status_label(event):
    return "Read" if event.get("lpr_read") else "Not Read"


def build_excel(events):
    wb = Workbook()
    ws = wb.active
    ws.title = "ANPR Events"

    headers = ["Plate Image"] + [label for _, label in EXPORT_COLUMNS] + ["LPR Status"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 14
    for col_index in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 16

    for row_index, event in enumerate(events, start=2):
        ws.row_dimensions[row_index].height = 62

        for col_offset, (field, _) in enumerate(EXPORT_COLUMNS, start=2):
            ws.cell(row=row_index, column=col_offset, value=event.get(field, ""))

        ws.cell(row=row_index, column=len(headers), value=lpr_status_label(event))

        image_path = resolve_image_path(event, "plate_image")
        thumb = make_thumbnail(image_path, EXPORT_THUMB_PX) if image_path else None

        if thumb:
            try:
                img = XLImage(thumb)
                img.width = EXPORT_THUMB_PX
                img.height = EXPORT_THUMB_PX
                ws.add_image(img, f"A{row_index}")
            except Exception:
                app.logger.warning("Failed to embed %s in Excel export", image_path)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def build_pdf(events):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph("ANPR Event Export", styles["Title"]),
        Spacer(1, 6),
    ]

    headers = ["Image"] + [label for _, label in EXPORT_COLUMNS] + ["LPR"]
    data = [headers]

    thumb_size = 16 * mm
    thumb_px = int(EXPORT_THUMB_PX)

    for event in events:
        image_path = resolve_image_path(event, "plate_image")
        thumb = make_thumbnail(image_path, thumb_px) if image_path else None

        if thumb:
            try:
                cell = RLImage(thumb, width=thumb_size, height=thumb_size)
            except Exception:
                cell = "N/A"
        else:
            cell = "N/A"

        row = [cell] + [str(event.get(field, "")) for field, _ in EXPORT_COLUMNS]
        row.append(lpr_status_label(event))
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2a44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fa")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return buffer


# ============================================================
# WEB ROUTES
# ============================================================

@app.route("/")
def dashboard():
    return render_template(
        "dashboard.html",
        poll_interval_ms=DASHBOARD_POLL_INTERVAL_MS,
    )


@app.route("/api/events")
def api_events():
    events, version = index.get_sorted_events()

    latest_entry = next(
        (event for event in events if event["direction"] == "Entry"),
        None,
    )

    latest_exit = next(
        (event for event in events if event["direction"] == "Exit"),
        None,
    )

    return jsonify({
        "events": events[:MAX_EVENTS_RETURNED],
        "latest_entry": latest_entry,
        "latest_exit": latest_exit,
        "stats": build_stats(events),
        "version": version,
    })


@app.route("/api/filters")
def api_filters():
    """Distinct vehicle_type/plate_type values seen in the index, so
    the frontend's export filter dropdowns reflect real data instead of
    a hardcoded guess."""

    events, _ = index.get_sorted_events()

    vehicle_types = sorted({
        event["vehicle_type"] for event in events
        if event["vehicle_type"] and event["vehicle_type"] != "-"
    })

    plate_types = sorted({
        event["plate_type"] for event in events
        if event["plate_type"] and event["plate_type"] != "-"
    })

    return jsonify({
        "vehicle_types": vehicle_types,
        "plate_types": plate_types,
    })


@app.route("/api/events/query")
def api_events_query():
    """Paginated, filtered event browsing - separate from /api/events
    (which stays a fixed-shape "latest 500" feed for the live dashboard)
    so browsing beyond that cap, or by date range/vehicle/plate type,
    doesn't require changing that existing contract."""

    events, version = index.get_sorted_events()
    filtered = apply_filters(events, request.args)

    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1

    try:
        page_size = int(request.args.get("page_size", 100))
    except ValueError:
        page_size = 100

    page_size = max(1, min(page_size, 1000))

    total = len(filtered)
    total_pages = max(1, -(-total // page_size))
    page = min(page, total_pages)

    start = (page - 1) * page_size
    end = start + page_size

    return jsonify({
        "events": filtered[start:end],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "version": version,
    })


@app.route("/api/export")
def api_export():
    fmt = (request.args.get("format") or "xlsx").lower()

    if fmt not in ("xlsx", "pdf"):
        abort(400, description="format must be 'xlsx' or 'pdf'")

    events, _ = index.get_sorted_events()
    filtered = apply_filters(events, request.args)

    if not filtered:
        abort(404, description="No events match the given filters")

    if len(filtered) > MAX_EXPORT_ROWS:
        abort(
            413,
            description=(
                f"{len(filtered)} events matched - narrow your date "
                f"range or filters to {MAX_EXPORT_ROWS} or fewer for "
                "an image-embedded export."
            ),
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if fmt == "xlsx":
        buffer = build_excel(filtered)
        return send_file(
            buffer,
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=f"anpr_export_{timestamp}.xlsx",
        )

    buffer = build_pdf(filtered)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"anpr_export_{timestamp}.pdf",
    )


@app.route("/api/events/stream")
def api_events_stream():
    """Server-Sent Events channel. Pushes a small "something changed"
    signal whenever the index updates - the client reacts by
    refetching /api/events, so the response payload shape/contract
    used everywhere else never has to be duplicated here."""

    def generate():
        wake = index.subscribe()

        try:
            yield "retry: 2000\n"
            yield f"event: update\ndata: {index.version}\n\n"

            while True:
                triggered = wake.wait(timeout=SSE_HEARTBEAT_SECONDS)

                if triggered:
                    wake.clear()
                    yield f"event: update\ndata: {index.version}\n\n"
                else:
                    yield ": heartbeat\n\n"

        except GeneratorExit:
            pass
        finally:
            index.unsubscribe(wake)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ============================================================
# IMAGE ROUTE
# ============================================================

@app.route("/images/<direction>/<path:filename>")
def image_file(direction, filename):
    if direction.lower() == "entry":
        directory = ENTRY_DIR
    elif direction.lower() == "exit":
        directory = EXIT_DIR
    else:
        abort(404)

    return send_from_directory(directory, filename)


# ============================================================
# HEALTH
# ============================================================

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "entry_directory": os.path.isdir(ENTRY_DIR),
        "exit_directory": os.path.isdir(EXIT_DIR),
        "indexed_events": len(index._events),
        "index_version": index.version,
    })


@app.route("/api/camera-status")
def api_camera_status():
    return jsonify({
        "entry": read_camera_status("Entry"),
        "exit": read_camera_status("Exit"),
    })


# ============================================================
# START
# ============================================================

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5020,
        debug=False,
        threaded=True,
    )
